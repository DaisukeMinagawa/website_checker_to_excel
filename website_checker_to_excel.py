import requests
from bs4 import BeautifulSoup
import time
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import difflib
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from dotenv import load_dotenv
import re
import pytz
import ssl
import logging

# ロギングの設定
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# .env ファイルから環境変数を読み込む
load_dotenv()

def get_website_content(url):
    """指定されたURLからウェブサイトの内容を取得する"""
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()  # HTTPエラーがある場合に例外を発生させる
        response.encoding = response.apparent_encoding  # 文字化け防止
        return response.text
    except requests.RequestException as e:
        logging.warning(f"ウェブサイトにアクセスできません: {url} - エラー: {e}")
        return None

def parse_html(content):
    """HTMLコンテンツを解析し、HTML構造とCSSを抽出する"""
    soup = BeautifulSoup(content, 'html.parser')
    html = soup.prettify()
    css = '\n'.join([style.string for style in soup.find_all('style')])
    return html, css

def get_diff(old_content, new_content):
    """2つのコンテンツの差分を取得する"""
    diff = difflib.unified_diff(old_content.splitlines(), new_content.splitlines(), lineterm='')
    return '\n'.join(list(diff)[2:])  # 最初の2行(差分ヘッダー)を除外

def write_to_excel(filename, timestamp, url, html_diff, css_diff):
    """変更内容をExcelファイルに書き込む"""
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "URL", "HTML Changes", "CSS Changes"])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    ws.append([timestamp, url, html_diff, css_diff])
    wb.save(filename)

def send_email(subject, body, to_email, attachment_path=None):
    """
    Excelファイルを添付してメールを送信する

    注意: Gmailを使用する場合、以下の手順が必要です
    1. Googleアカウントの2段階認証を有効にする
    2. アプリパスワードを生成する(Google アカウント > セキュリティ > アプリパスワード)
    3. 生成されたアプリパスワードを EMAIL_PASS 環境変数に設定する
    """
    from_email = os.getenv('EMAIL_USER')
    password = os.getenv('EMAIL_PASS')
    smtp_server = os.getenv('SMTP_SERVER')
    smtp_port = int(os.getenv('SMTP_PORT'))
    use_tls = os.getenv('USE_TLS', 'True').lower() == 'true'

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())

        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(attachment_path)}")
        msg.attach(part)

    try:
        if use_tls:
            context = ssl.create_default_context()
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls(context=context)
                server.login(from_email, password)
                server.send_message(msg)
        else:
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(from_email, password)
                server.send_message(msg)
        logging.info("メールが正常に送信されました")
    except Exception as e:
        logging.error(f"メール送信中にエラーが発生しました: {e}")

def get_valid_url():
    """有効なURLを取得する"""
    while True:
        url = input("監視するウェブサイトのURLを入力してください(httpとhttpsは省略可能): ").strip()
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url

        # 簡単な正規表現でURLの形式を確認
        if re.match(r'https?://[\w\-\.]+\.[a-zA-Z]{2,}', url):
            return url
        print("無効なURLです。正しいURLを入力してください。")

def get_japan_time():
    """現在の日本時間を取得する"""
    japan_tz = pytz.timezone('Asia/Tokyo')
    return datetime.now(japan_tz).strftime("%Y-%m-%d %H:%M:%S")

def main():
    url = get_valid_url()

    excel_filename = input("保存するExcelファイルの名前を入力してください(拡張子.xlsxは自動で追加されます): ")
    excel_filename += ".xlsx"

    to_email = os.getenv('TO_EMAIL')
    if not to_email:
        logging.error("エラー: TO_EMAILが設定されていません。.envファイルを確認してください。")
        return

    logging.info(f"{url} の監視を開始します。変更は {excel_filename} に記録され、設定されたメールアドレスに送信されます。")

    previous_content = None
    site_available = False
    first_successful_access = None

    while True:
        current_content = get_website_content(url)

        if current_content is None:
            if not site_available:
                logging.info(f"サイト {url} はまだ利用できません。監視を続けます...")
            else:
                logging.warning(f"サイト {url} にアクセスできません。一時的なエラーの可能性があります。")
            time.sleep(300)  # 5分後に再試行
            continue

        if not site_available:
            site_available = True
            first_successful_access = get_japan_time()
            logging.info(f"サイト {url} が利用可能になりました。最初のアクセス時刻: {first_successful_access}")
            send_email(
                subject=f"ウェブサイト監視開始 - {url}",
                body=f"ウェブサイト {url} が利用可能になりました。監視を開始します。\n最初のアクセス時刻: {first_successful_access}",
                to_email=to_email
            )
            previous_content = current_content
            previous_html, previous_css = parse_html(current_content)
            continue

        current_html, current_css = parse_html(current_content)

        html_diff = get_diff(previous_html, current_html)
        css_diff = get_diff(previous_css, current_css)

        if html_diff or css_diff:
            timestamp = get_japan_time()
            write_to_excel(excel_filename, timestamp, url, html_diff, css_diff)
            logging.info(f"変更を検出しました。{timestamp} に記録しました。")

            # メール送信
            subject = f"ウェブサイト変更通知 - {url}"
            body = f"ウェブサイト {url} に変更が検出されました。詳細は添付のExcelファイルをご確認ください。"
            send_email(subject, body, to_email, excel_filename)

            previous_html, previous_css = current_html, current_css

        time.sleep(1800)  # 30分(1800秒)待機

if __name__ == "__main__":
    main()
